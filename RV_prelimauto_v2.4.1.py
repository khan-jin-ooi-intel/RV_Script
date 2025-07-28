import pandas as pd
import argparse
import shutil
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Font, PatternFill

# Load data from Excel file
def load_excel(file_path):
    #return pd.read_excel(file_path)
    return pd.read_csv(file_path)

# Split Dataset (df - Dataframe) into 1st test and Retest
def retest_check(df_latest, location_df):
    df_retest = pd.Series([])
    # Didnt run through this location, hence empty dataframe
    if df_latest.size == 0:
        df_first = pd.Series([])
    # 1st Test == Latest Run
    elif df_latest["Within LOTS Sequence Num"].iloc[0] == 1:
        df_first = df_latest
    # 1st Test != Latest Run
    else:
        df_retest = df_latest
        df_first = location_df.loc[location_df["Within LOTS Sequence Num"] == 1]
    return df_first, df_retest

# Filter and Extract data from "TEST_RESULT" column based on keywords and filter_keywords
def result_extract(df, column="", key="", keywords=[], filter_keywords=[], use_regex=False):
    if use_regex == False:
        if df.size == 0:
            data = "-"
        else:
            data = df[column].iloc[0]
    elif use_regex == True:
        data = "-"
        # match by keywords only
        if filter_keywords is None:
            pattern = "(?=" + ")(?=.*".join(keywords) + ")"
        # match by keywords and filter out filter_keywords
        else:
            pattern =  "(?!.*(" + "|".join(filter_keywords) + "))(?=" + ")(?=.*".join(keywords) + ")"
        filtered_df = df[df["TEST_NAME"].str.contains(pattern, case=True, regex=True)]
        # Data valid if there is only a single match
        if filtered_df.shape[0] == 1:
            data = filtered_df["TEST_RESULT"].iloc[0]
        elif filtered_df.shape[0] > 1:
            data = "Duplicates Found, Please Optimize Keywords"
            print("="*100)
            print(f"Duplicates Found for token [{key}], Please Optimize Keywords")
            print("="*100)
            #pd.set_option('display.max_columns',None)
            pd.set_option('display.max_colwidth', None)
            print(filtered_df["TEST_NAME"])
            print("="*100)
        # Remain only VMIN (ex. 0.780|0.760|1.100|2)
        if 'vmin' in key:
            data = data.split('|')[0]
    return data

# Extract all the data as per defined in the dictionaries (sort_info, class_info, etc.)
def data_pull(df_info, df_first, df_retest, df_latest, results_arr, socket):
    # Append Location Here
    socket_name = {
        '119325'    :   "sort_",
        '6261'      :   "classhot_",
        '6212'      :   "classcold_",
        '5242'      :   "qahot_",
        '5243'      :   "qacold_",
    }

    # Check if Socket Exist for this script
    try:
        socket_tp = result_extract(df_latest, column="Program Name")
        socket_bin = result_extract(df_latest, column="FUNCTIONAL_BIN")
        socket_lot, socket_hvqk = "-","-"
        if socket != '119325':
            socket_lot = result_extract(df_latest, column="Lot")
        else:
            socket_hvqk = "True" if not df_latest[df_latest["TEST_NAME"].str.contains("_POSTHVQK", case=True, regex=True)].empty else "False"
            
        results_arr.update({
            socket_name[socket]+"tprev":socket_tp,
            socket_name[socket]+"bin":socket_bin,
            socket_name[socket]+"lot":socket_lot,
            socket_name[socket]+"hvqk":socket_hvqk,
        })
    
        # key -> "sort_desku...", sub_key -> "default","keywords","filter_keywords"
        for key, sub_dict in df_info.items():
            # retest results present
            if "retest" in key and df_retest.size != 0:
                data = {socket_name[socket]+key: result_extract(df_retest, key=key, keywords=sub_dict['Keywords'], filter_keywords=sub_dict['Exclude_Keywords'], use_regex=True)}
            # extract non-retest results
            elif "retest" not in key and df_first.size != 0:
                data = {socket_name[socket]+key: result_extract(df_first, key=key, keywords=sub_dict['Keywords'], filter_keywords=sub_dict['Exclude_Keywords'], use_regex=True)}
            # retest results not present
            else:
                data = {socket_name[socket]+key: sub_dict['Default_Value']}
            results_arr.update(data)
    except:
        print(f"Socket not defined/enabled for this script, skipping location [{socket}]...")

    return results_arr

# Extract all data revelant to a socket (SORT, CLASSHOT, QAHOT, etc.)
def data_compile(rawdata_df, VID, locn_list, sort_info, class_info):
    # Remain only Data related to VID
    filtered_df = rawdata_df.loc[rawdata_df["VISUAL_ID"] == VID]
    # Extract General Data
    if filtered_df.size != 0:
        sort_ult = "_".join(filtered_df[["SORT_LOT","SORT_WAFER","SORT_X","SORT_Y"]].iloc[0].astype(str))
        sort_smp = filtered_df.query('TEST_NAME == "DFF_SMPULVT"')["TEST_RESULT"].iat[0]
        burn_in = "True" if any(x == 7652 for x in filtered_df["Operation"].unique()) else "False"
    
    # Extract Sort/Class Specific Data
    results_arr = {}
    for locn in locn_list:
        # Filter by Location
        socket_df = filtered_df.loc[filtered_df["Operation"] == int(locn)]
        # Remain Most Recent Lot
        socket_lot = socket_df.sort_values(by="LOTS End Date Time", ascending=False)
        socket_df_latest = socket_df.loc[(socket_df["Lot"] == socket_lot.iloc[0]["Lot"])] if socket_lot.size != 0 else socket_df
        # Define Initial (First) and Retest 
        socket_df_first, socket_df_retest = retest_check(socket_df_latest.loc[socket_df_latest["Operation Final Latest Flag"] == "Y"], socket_df_latest)
        # Pull Data
        if locn != '119325':
            socket_results = data_pull(class_info, socket_df_first, socket_df_retest, socket_df_latest, results_arr, socket=locn)
        else:
            socket_results = data_pull(sort_info, socket_df_first, socket_df_retest, socket_df_latest, results_arr, socket=locn)
             
    # Updated dictionaries containing all the extracted results
    unit_results = {
        'VID': VID,
        'sort_ult': sort_ult,
        'sort_smp': sort_smp,
        'burn_in': burn_in
    }
    # Concatenate into single dictionary
    results = unit_results | socket_results
    
    return results

def replace_n_write(df, worksheet, table_startrow, table_columns):
    start_col = column_index_from_string(table_columns.split(":")[0])
    end_col = column_index_from_string(table_columns.split(":")[1])
    for col in range(start_col,end_col+1):
        for index, value in enumerate(df.iloc[:,col-start_col], start=table_startrow):
            worksheet.cell(row=index, column=col, value=value)

# Save filtered data to a new Excel file
def save_to_excel(results, input_format, output_filepath, VID, token_dump):
    
    # If Output file Exists
    try:
        wb = load_workbook(output_filepath)
    # If Output file does not Exists
    except:
        # Create empty workbook and remove the default empty sheet
        shutil.copy(input_format, output_filepath)
        wb = load_workbook(output_filepath)
    
    # Duplicate Sample Output Sheet
    sample_sheet = wb["sample"]
    clone_sheet = wb.copy_worksheet(sample_sheet)
    clone_sheet.title = VID

    tables_df = pd.read_excel(input_format,sheet_name="table_params")
    tables_dict = tables_df.set_index("Table").to_dict(orient='index')
    
    for table, params in tables_dict.items():
        data_columns = params['Columns']
        startrow = params['StartRow']
        endrow = params['EndRow']
        # Read data table from Excel
        df = pd.read_excel(input_format,sheet_name='sample',na_filter=False,header=None,
                           usecols=data_columns,
                           skiprows=startrow-1,
                           nrows=endrow-startrow+1)
        # Replace table with extracted data
        df.replace(results,inplace=True)
        # Print to output excel
        replace_n_write(df,wb[VID],startrow,data_columns)
   
    # Dump all extracted tokens into another sheet
    if token_dump:
        ws = wb.create_sheet(title=f"{VID[-3:]}_tokens")
        # Insert header
        ws.append(['Token','Value'])
        for token, value in results.items():
            ws.append([token,value])

    wb.save(output_filepath)

def adjust_width(ws, width):
    # Adjust the width of all columns
    for col in ws.iter_cols():
        col_letter = col[0].column_letter  # Get column letter
        ws.column_dimensions[col_letter].width = width

def adjust_cells(output_file_path):
    # Load the workbook
    wb = load_workbook(output_file_path)
    input_sheets = ['sort_tokens','class_tokens','sample','table_params','compare']
    for sheet in input_sheets:
        wb.remove(wb[sheet])
 
    # Loop through all sheets
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        
        if "token" not in sheet:
            adjust_width(ws,width=16)

            # Compare results within defined lists and highlight for any mismatch    
            compare_df = pd.read_excel(input_format,sheet_name="compare")
            compare_dict = compare_df.set_index("List").to_dict(orient='index')
            for list, cell_list in compare_dict.items():
                # Remove blanks
                cell_value_map = {cell:ws[cell].value for cell in cell_list['Cells'].split(",") if ws[cell].value != '-'}
                # Remain Unique Values
                duplicates = Counter(cell_value_map.values())
                if duplicates:
                    max_occurs = max(duplicates.values())
                    unique_cells = {cell:value for cell, value in cell_value_map.items() if duplicates[value] != max_occurs}
                    for cell in unique_cells:
                        ws[cell].fill = PatternFill(start_color="00FFFF00", fill_type="solid")

        else:
            adjust_width(ws,width=33)
            title_cells = ['A1','B1']
            for cell in title_cells:
                c = ws[cell]
                c.fill = PatternFill(start_color="000000FF", fill_type="solid")
                c.font = Font(size=15,bold=True,color="00FFFFFF")

    # Save the modified workbook
    wb.save(output_file_path)

def main(input_filepath, input_format, output_filepath, vid_list, locn_list, token_dump):
    # Load all Input Files
    rawdata_df = pd.read_csv(input_filepath).fillna("")
    sort_df = pd.read_excel(input_format, sheet_name="sort_tokens")
    class_df = pd.read_excel(input_format, sheet_name="class_tokens")
    # Converting column values to list for later operation
    columns_to_list = ['Keywords', 'Exclude_Keywords']
    sort_df[columns_to_list] = sort_df[columns_to_list].map(lambda x: x.rsplit(",") if isinstance(x,str) else None)
    class_df[columns_to_list] = class_df[columns_to_list].map(lambda x: x.rsplit(",") if isinstance(x,str) else None)
    # Convert dataframe to dictionary
    sort_info = sort_df.set_index("Variable").to_dict(orient="index")
    class_info = class_df.set_index("Variable").to_dict(orient="index")
    # Loop through each unit for data pulling
    for unit in vid_list:
        results = data_compile(rawdata_df, unit, locn_list, sort_info, class_info)
        save_to_excel(results, input_format, output_filepath, unit, token_dump)
        print(f"Finished Extracting Data for VID [{unit}]"+"\n"+"="*100)
    print("Data for All Units Extracted Successfully...\n")

    adjust_cells(output_filepath)
    print(f"Cells highlighted, fitted and saved to '{output_filepath}'\n")
    
if __name__ == "__main__":

    parser = argparse.ArgumentParser(description='This script extracts specified data from input csv file for given VIDs')
    parser.add_argument("--inputfile", type=str, required=True, help="Input file name (include .csv)")
    parser.add_argument("--outputfile", type=str, help="Output file name (include .xlsx)")
    parser.add_argument("--format", type=str, required=True, help="Output file format (include .xlsx)")
    parser.add_argument("--vid", type=str, required=True, help='List of VIDs seperated by commas, ex. U4F771Q300179,U4F771Q300180')
    parser.add_argument("--locn", type=str, required=True, help='List of Locations required seperated by commas, ex. 119325,6261,6212')
    parser.add_argument("--dump", action='store_true', help='Dumps all extracted tokens into seperate excel sheet')

    args = parser.parse_args()

    # Input and Output file paths
    input_filepath = Path(args.inputfile)
    input_format = Path(args.format)
    vid_list = args.vid.split(",")
    locn_list = args.locn.split(",")
    token_dump = args.dump

    if args.outputfile:
        output_filepath = Path(args.outputfile)
    else:
        output_filepath = input_filepath.parent.joinpath("PROCESSED_RESULTS.xlsx")
        print(f"Output filepath not defined. Defaulting to input filepath...")
    
    print("="*100)
    print(f"INPUT FILE PATH: \t{input_filepath}")
    print(f"OUTPUT FILE PATH: \t{output_filepath}")
    print(f"FORMAT PATH: \t\t{input_format}")
    print(f"VID List: \t\t{vid_list}")
    print("="*100)
    
    if output_filepath.is_file():
        output_filepath.unlink()
        print("Output file found in provided path, File Removed...\n")
    
    if not (input_filepath.exists() and input_format.exists()) :
        print("Input Parameters not complete, aborting script...\n")
    else: 
        main(input_filepath, input_format, output_filepath, vid_list, locn_list, token_dump)