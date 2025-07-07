import pandas as pd
import argparse
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Side, Border

# Load data from Excel file
def load_excel(file_path):
    #return pd.read_excel(file_path)
    return pd.read_csv(file_path)

# Split Dataset (df - Dataframe) into 1st test and Retest
def retest_check(df_latest, location_df):
    df_retest = pd.Series([])
    # Didnt run through this location, hence empty dataframe
    if df_latest.size == 0:
        df_first = pd.Series([])
    # 1st Test == Latest Run
    elif df_latest["Within LOTS Sequence Num"].iloc[0] == 1:
        df_first = df_latest
    # 1st Test != Latest Run
    else:
        df_retest = df_latest
        df_first = location_df.loc[location_df["Within LOTS Sequence Num"] == 1]
    return df_first, df_retest

# Filter and Extract data from "TEST_RESULT" column based on keywords and filter_keywords
def result_extract(df, column="", key="", keywords=[], filter_keywords=[], use_regex=False):
    if use_regex == False:
        if df.size == 0:
            data = "-"
        else:
            data = df[column].iloc[0]
    elif use_regex == True:
        data = "-"
        # match by keywords only
        if filter_keywords is None:
            pattern = "(?=" + ")(?=.*".join(keywords) + ")"
        # match by keywords and filter out filter_keywords
        else:
            pattern =  "(?!.*(" + "|".join(filter_keywords) + "))(?=" + ")(?=.*".join(keywords) + ")"
        filtered_df = df[df["TEST_NAME"].str.contains(pattern, case=True, regex=True)]
        # Data valid if there is only a single match
        if filtered_df.shape[0] == 1:
            data = filtered_df["TEST_RESULT"].iloc[0]
        elif filtered_df.shape[0] > 1:
            data = "Duplicates Found, Please Optimize Keywords"
            print("="*100)
            print("Duplicates Found, Please Optimize Keywords")
            print("="*100)
            #pd.set_option('display.max_columns',None)
            pd.set_option('display.max_colwidth', None)
            print(filtered_df["TEST_NAME"])
        # Remain only VMIN (ex. 0.780|0.760|1.100|2)
        if 'vmin' in key:
            data = data.split('|')[0]
    return data

# Extract all the data as per defined in the dictionaries (sort_info, class_info, etc.)
def data_pull(df_info, df_first, df_retest, name=""):
    results_arr = {}
    # key -> "sort_desku...", sub_key -> "default","keywords","filter_keywords"
    for key, sub_dict in df_info.items():
        # retest results present
        if "retest" in key and df_retest.size != 0:
            data = {name+key: result_extract(df_retest, key=key, keywords=sub_dict['Keywords'], filter_keywords=sub_dict['Exclude_Keywords'], use_regex=True)}
        # extract non-retest results
        elif "retest" not in key and df_first.size != 0:
            data = {name+key: result_extract(df_first, key=key, keywords=sub_dict['Keywords'], filter_keywords=sub_dict['Exclude_Keywords'], use_regex=True)}
        # retest results not present
        else:
            data = {name+key: sub_dict['Default_Value']}
        results_arr.update(data)
    return results_arr

def data_compile(rawdata_df, VID, sort_info, class_info):
    # Remain only Data related to VID
    filtered_df = rawdata_df.loc[rawdata_df["VISUAL_ID"] == VID]
    # Filter by Location
    sort_df = filtered_df.loc[filtered_df["Operation"] == 119325]
    classhot_df = filtered_df.loc[filtered_df["Operation"] == 6261]
    classcold_df = filtered_df.loc[filtered_df["Operation"] == 6212]
    # Remain Most Recent Lot at each location
    sort_lot = sort_df.sort_values(by="LOTS End Date Time", ascending=False)
    classhot_lot = classhot_df.sort_values(by="LOTS End Date Time", ascending=False)
    classcold_lot = classcold_df.sort_values(by="LOTS End Date Time", ascending=False)
    sort_df_latest = sort_df.loc[(sort_df["Lot"] == sort_lot.iloc[0]["Lot"])] if sort_lot.size != 0 else sort_df
    classhot_df_latest = classhot_df.loc[(classhot_df["Lot"] == classhot_lot.iloc[0]["Lot"])] if classhot_lot.size != 0 else classhot_df
    classcold_df_latest = classcold_df.loc[(classcold_df["Lot"] == classcold_lot.iloc[0]["Lot"])] if classcold_lot.size != 0 else classcold_df
    # Define Initial (First) and Retest 
    sort_df_first, sort_df_retest = retest_check(sort_df_latest.loc[sort_df_latest["Operation Final Latest Flag"] == "Y"], sort_df_latest)
    classhot_df_first, classhot_df_retest = retest_check(classhot_df_latest.loc[classhot_df_latest["Operation Final Latest Flag"] == "Y"],classhot_df_latest)
    classcold_df_first, classcold_df_retest = retest_check(classcold_df_latest.loc[classcold_df_latest["Operation Final Latest Flag"] == "Y"],classcold_df_latest)

    # Extract Data
    if filtered_df.size != 0:
        sort_ult = "_".join(filtered_df[["SORT_LOT","SORT_WAFER","SORT_X","SORT_Y"]].iloc[0].astype(str))
        sort_smp = filtered_df.query('TEST_NAME == "DFF_SMPULVT"')["TEST_RESULT"].iat[0]
        burn_in = True if any(x == 7652 for x in filtered_df["Operation"].unique()) else False
        sort_tp, classhot_tp, classcold_tp = result_extract(sort_df_latest, column="Program Name"), result_extract(classhot_df_latest, column="Program Name"), result_extract(classcold_df_latest, column="Program Name")
        sort_bin, classhot_bin, classcold_bin = result_extract(sort_df_latest, column="FUNCTIONAL_BIN"), result_extract(classhot_df_latest, column="FUNCTIONAL_BIN"), result_extract(classcold_df_latest, column="FUNCTIONAL_BIN")
        classhot_lot, classcold_lot = result_extract(classhot_df_latest, column="Lot"), result_extract(classcold_df_latest, column="Lot")
        sort_hvqk = True if not sort_df_latest[sort_df_latest["TEST_NAME"].str.contains("_POSTHVQK", case=True, regex=True)].empty else False
    else:
        sort_ult, sort_smp, burn_in, sort_tp, classhot_tp, classcold_tp, sort_bin, classhot_bin, classcold_bin, classhot_lot, classcold_lot, sort_hvqk = "-","-","-","-","-","-","-","-","-","-","-","-"

    # Updated dictionaries containing all the extracted results
    unit_results = {
        'VID': VID,
        'SORT_ULT': sort_ult,
        'SORT_SMP': sort_smp,
        'SORT_TPREV': sort_tp, 
        'SORT_HVQK': sort_hvqk,
        'SORT_BIN': sort_bin,
        'CLASSHOT_LOT': classhot_lot,
        'CLASSHOT_TPREV': classhot_tp,
        'CLASSHOT_BIN': classhot_bin,
        'CLASSCOLD_LOT': classcold_lot,
        'CLASSCOLD_TPREV': classcold_tp,
        'CLASSCOLD_BIN': classcold_bin,
        'BURN_IN': burn_in
    }
    sort_results = data_pull(sort_info, sort_df_first, sort_df_retest)
    classhot_results = data_pull(class_info, classhot_df_first, classhot_df_retest, name="classhot_")
    classcold_results = data_pull(class_info, classcold_df_first, classcold_df_retest, name="classcold_")

    # Compile Extracted Data, build dataframes accordinly to desired table format
    unit_table = (
        [unit_results['SORT_ULT'],unit_results['CLASSHOT_LOT'],unit_results['CLASSCOLD_LOT']],              # ULT / LOT
        [unit_results['SORT_TPREV'],unit_results['CLASSHOT_TPREV'],unit_results['CLASSCOLD_TPREV']],        # TP REV
        [unit_results['SORT_BIN'],unit_results['CLASSHOT_BIN'],unit_results['CLASSCOLD_BIN']],              # BINNING
        [unit_results['SORT_SMP'],'',''],                                                                   # SMP
        [unit_results['BURN_IN'],'','']                                                                     # BURN IN
    )
    de_table = (
        [sort_results['sort_desku_begin_first'],sort_results['sort_desku_posthvqk_first'],sort_results['sort_desku_end_trad_first'],sort_results['sort_desku_end_ca_first'],sort_results['sort_desku_tpi_end_first'],sort_results['sort_desku_tpi_updated_end_first'],classhot_results['classhot_desku_trad_first'],classhot_results['classhot_desku_ca_first'],classhot_results['classhot_desku_tpi_first'],classhot_results['classhot_desku_tpi_updated_first'],classcold_results['classcold_desku_tpi_first'],classcold_results['classcold_desku_tpi_updated_first']],
        [sort_results['sort_desku_begin_retest'],sort_results['sort_desku_posthvqk_retest'],sort_results['sort_desku_end_trad_retest'],sort_results['sort_desku_end_ca_retest'],sort_results['sort_desku_tpi_end_retest'],sort_results['sort_desku_tpi_updated_end_retest'],classhot_results['classhot_desku_trad_retest'],classhot_results['classhot_desku_ca_retest'],classhot_results['classhot_desku_tpi_retest'],classhot_results['classhot_desku_tpi_updated_retest'],classcold_results['classcold_desku_tpi_retest'],classcold_results['classcold_desku_tpi_updated_retest']],
        ['','','','','','','','','','','',''],
        [sort_results['sort_defeid_begin_first'],sort_results['sort_defeid_posthvqk_first'],sort_results['sort_defeid_end_trad_first'],sort_results['sort_defeid_end_ca_first'],sort_results['sort_defeid_tpi_end_first'],sort_results['sort_defeid_tpi_updated_end_first'],classhot_results['classhot_defeid_trad_first'],classhot_results['classhot_defeid_ca_first'],classhot_results['classhot_defeid_tpi_first'],classhot_results['classhot_defeid_tpi_updated_first'],classcold_results['classcold_defeid_tpi_first'],classcold_results['classcold_defeid_tpi_updated_first']],
        [sort_results['sort_defeid_begin_retest'],sort_results['sort_defeid_posthvqk_retest'],sort_results['sort_defeid_end_trad_retest'],sort_results['sort_defeid_end_ca_retest'],sort_results['sort_defeid_tpi_end_retest'],sort_results['sort_defeid_tpi_updated_end_retest'],classhot_results['classhot_defeid_trad_retest'],classhot_results['classhot_defeid_ca_retest'],classhot_results['classhot_defeid_tpi_retest'],classhot_results['classhot_defeid_tpi_updated_retest'],classcold_results['classcold_defeid_tpi_retest'],classcold_results['classcold_defeid_tpi_updated_retest']]
    )
    gt_table = (
        [sort_results['sort_gtsku_begin_first'],sort_results['sort_gtsku_posthvqk_first'],sort_results['sort_gtsku_end_trad_first'],sort_results['sort_gtsku_end_ca_first'],sort_results['sort_gtsku_end_isol_first'],sort_results['sort_gtsku_tpi_end_first'],sort_results['sort_gtsku_tpi_updated_end_first'],classhot_results['classhot_gtsku_trad_first'],classhot_results['classhot_gtsku_ca_first'],classhot_results['classhot_gtsku_isol_first'],classhot_results['classhot_gtsku_tpi_first'],classhot_results['classhot_gtsku_tpi_updated_first'],classcold_results['classcold_gtsku_trad_first'],classcold_results['classcold_gtsku_ca_first'],classcold_results['classcold_gtsku_isol_first']],
        [sort_results['sort_gtsku_begin_retest'],sort_results['sort_gtsku_posthvqk_retest'],sort_results['sort_gtsku_end_trad_retest'],sort_results['sort_gtsku_end_ca_retest'],sort_results['sort_gtsku_end_isol_retest'],sort_results['sort_gtsku_tpi_end_retest'],sort_results['sort_gtsku_tpi_updated_end_retest'],classhot_results['classhot_gtsku_trad_retest'],classhot_results['classhot_gtsku_ca_retest'],classhot_results['classhot_gtsku_isol_retest'],classhot_results['classhot_gtsku_tpi_retest'],classhot_results['classhot_gtsku_tpi_updated_retest'],classcold_results['classcold_gtsku_trad_retest'],classcold_results['classcold_gtsku_ca_retest'],classcold_results['classcold_gtsku_isol_retest']],
        ['','','','','','','','','','','','','','',''],
        [sort_results['sort_gtfrg_begin_first'],sort_results['sort_gtfrg_posthvqk_first'],sort_results['sort_gtfrg_end_trad_first'],sort_results['sort_gtfrg_end_ca_first'],sort_results['sort_gtfrg_end_isol_first'],sort_results['sort_gtfrg_tpi_end_first'],sort_results['sort_gtfrg_tpi_updated_end_first'],classhot_results['classhot_gtfrg_trad_first'],classhot_results['classhot_gtfrg_ca_first'],classhot_results['classhot_gtfrg_isol_first'],classhot_results['classhot_gtfrg_tpi_first'],classhot_results['classhot_gtfrg_tpi_updated_first'],classcold_results['classcold_gtfrg_trad_first'],classcold_results['classcold_gtfrg_ca_first'],classcold_results['classcold_gtfrg_isol_first']],
        [sort_results['sort_gtfrg_begin_retest'],sort_results['sort_gtfrg_posthvqk_retest'],sort_results['sort_gtfrg_end_trad_retest'],sort_results['sort_gtfrg_end_ca_retest'],sort_results['sort_gtfrg_end_isol_retest'],sort_results['sort_gtfrg_tpi_end_retest'],sort_results['sort_gtfrg_tpi_updated_end_retest'],classhot_results['classhot_gtfrg_trad_retest'],classhot_results['classhot_gtfrg_ca_retest'],classhot_results['classhot_gtfrg_isol_retest'],classhot_results['classhot_gtfrg_tpi_retest'],classhot_results['classhot_gtfrg_tpi_updated_retest'],classcold_results['classcold_gtfrg_trad_retest'],classcold_results['classcold_gtfrg_ca_retest'],classcold_results['classcold_gtfrg_isol_retest']],
        ['','','','','','','','','','','','','','',''],
        [sort_results['sort_gtfeid_begin_first'],sort_results['sort_gtfeid_posthvqk_first'],sort_results['sort_gtfeid_end_trad_first'],sort_results['sort_gtfeid_end_ca_first'],sort_results['sort_gtfeid_end_isol_first'],sort_results['sort_gtfeid_tpi_end_first'],sort_results['sort_gtfeid_tpi_updated_end_first'],classhot_results['classhot_gtfeid_trad_first'],classhot_results['classhot_gtfeid_ca_first'],classhot_results['classhot_gtfeid_isol_first'],classhot_results['classhot_gtfeid_tpi_first'],classhot_results['classhot_gtfeid_tpi_updated_first'],classcold_results['classcold_gtfeid_trad_first'],classcold_results['classcold_gtfeid_ca_first'],classcold_results['classcold_gtfeid_isol_first']],
        [sort_results['sort_gtfeid_begin_retest'],sort_results['sort_gtfeid_posthvqk_retest'],sort_results['sort_gtfeid_end_trad_retest'],sort_results['sort_gtfeid_end_ca_retest'],sort_results['sort_gtfeid_end_isol_retest'],sort_results['sort_gtfeid_tpi_end_retest'],sort_results['sort_gtfeid_tpi_updated_end_retest'],classhot_results['classhot_gtfeid_trad_retest'],classhot_results['classhot_gtfeid_ca_retest'],classhot_results['classhot_gtfeid_isol_retest'],classhot_results['classhot_gtfeid_tpi_retest'],classhot_results['classhot_gtfeid_tpi_updated_retest'],classcold_results['classcold_gtfeid_trad_retest'],classcold_results['classcold_gtfeid_ca_retest'],classcold_results['classcold_gtfeid_isol_retest']]
    )
    media_table = (
        [600,sort_results['sort_media_f1_vmin_first'],sort_results['sort_media_f1_vmin_retest'],classhot_results['classhot_media_f1_srh_vmin_first'],classhot_results['classhot_media_f1_srh_vmin_retest'],classhot_results['classhot_media_f1_chk_vmin_first'],classhot_results['classhot_media_f1_chk_vmin_retest'],classcold_results['classcold_media_f1_chk_vmin_first'],classcold_results['classcold_media_f1_chk_vmin_retest']],
        [1100,sort_results['sort_media_f2_vmin_first'],sort_results['sort_media_f2_vmin_retest'],classhot_results['classhot_media_f2_srh_vmin_first'],classhot_results['classhot_media_f2_srh_vmin_retest'],classhot_results['classhot_media_f2_chk_vmin_first'],classhot_results['classhot_media_f2_chk_vmin_retest'],classcold_results['classcold_media_f2_chk_vmin_first'],classcold_results['classcold_media_f2_chk_vmin_retest']],
        [1500,sort_results['sort_media_f3_vmin_first'],sort_results['sort_media_f3_vmin_retest'],classhot_results['classhot_media_f3_srh_vmin_first'],classhot_results['classhot_media_f3_srh_vmin_retest'],classhot_results['classhot_media_f3_chk_vmin_first'],classhot_results['classhot_media_f3_chk_vmin_retest'],classcold_results['classcold_media_f3_chk_vmin_first'],classcold_results['classcold_media_f3_chk_vmin_retest']],
        ['','','','','','','','',''],
        [600,sort_results['sort_media_f1_lp_first'],sort_results['sort_media_f1_lp_retest'],classhot_results['classhot_media_f1_srh_lp_first'],classhot_results['classhot_media_f1_srh_lp_retest'],classhot_results['classhot_media_f1_chk_lp_first'],classhot_results['classhot_media_f1_chk_lp_retest'],classcold_results['classcold_media_f1_chk_lp_first'],classcold_results['classcold_media_f1_chk_lp_retest']],
        [1100,sort_results['sort_media_f2_lp_first'],sort_results['sort_media_f2_lp_retest'],classhot_results['classhot_media_f2_srh_lp_first'],classhot_results['classhot_media_f2_srh_lp_retest'],classhot_results['classhot_media_f2_chk_lp_first'],classhot_results['classhot_media_f2_chk_lp_retest'],classcold_results['classcold_media_f2_chk_lp_first'],classcold_results['classcold_media_f2_chk_lp_retest']],
        [1500,sort_results['sort_media_f3_lp_first'],sort_results['sort_media_f3_lp_retest'],classhot_results['classhot_media_f3_srh_lp_first'],classhot_results['classhot_media_f3_srh_lp_retest'],classhot_results['classhot_media_f3_chk_lp_first'],classhot_results['classhot_media_f3_chk_lp_retest'],classcold_results['classcold_media_f3_chk_lp_first'],classcold_results['classcold_media_f3_chk_lp_retest']]
    )
    render_table = (
        [600,sort_results['sort_render_f1_vmin_first'],sort_results['sort_render_f1_vmin_retest'],classhot_results['classhot_render_f1_srh_vmin_first'],classhot_results['classhot_render_f1_srh_vmin_retest'],classhot_results['classhot_render_f1_chk_vmin_first'],classhot_results['classhot_render_f1_chk_vmin_retest'],classcold_results['classcold_render_f1_chk_vmin_first'],classcold_results['classcold_render_f1_chk_vmin_retest']],
        [1600,'NA','NA',classhot_results['classhot_render_f2_srh_vmin_first'],classhot_results['classhot_render_f2_srh_vmin_retest'],'NA','NA','NA','NA'],
        [2000,sort_results['sort_render_f3_vmin_first'],sort_results['sort_render_f3_vmin_retest'],classhot_results['classhot_render_f3_srh_vmin_first'],classhot_results['classhot_render_f3_srh_vmin_retest'],classhot_results['classhot_render_f3_chk_vmin_first'],classhot_results['classhot_render_f3_chk_vmin_retest'],classcold_results['classcold_render_f3_chk_vmin_first'],classcold_results['classcold_render_f3_chk_vmin_retest']],
        [2400,'NA','NA',classhot_results['classhot_render_f4_srh_vmin_first'],classhot_results['classhot_render_f4_srh_vmin_retest'],'NA','NA','NA','NA'],
        [2900,sort_results['sort_render_f5_vmin_first'],sort_results['sort_render_f5_vmin_retest'],classhot_results['classhot_render_f5_srh_vmin_first'],classhot_results['classhot_render_f5_srh_vmin_retest'],classhot_results['classhot_render_f5_chk_vmin_first'],classhot_results['classhot_render_f5_chk_vmin_retest'],classcold_results['classcold_render_f5_chk_vmin_first'],classcold_results['classcold_render_f5_chk_vmin_retest']],
        ['','','','','','','','',''],
        [600,sort_results['sort_render_f1_lp_first'],sort_results['sort_render_f1_lp_retest'],classhot_results['classhot_render_f1_srh_lp_first'],classhot_results['classhot_render_f1_srh_lp_retest'],classhot_results['classhot_render_f1_chk_lp_first'],classhot_results['classhot_render_f1_chk_lp_retest'],classcold_results['classcold_render_f1_chk_lp_first'],classcold_results['classcold_render_f1_chk_lp_retest']],
        [1600,'NA','NA',classhot_results['classhot_render_f2_srh_lp_first'],classhot_results['classhot_render_f2_srh_lp_retest'],'NA','NA','NA','NA'],
        [2000,sort_results['sort_render_f3_lp_first'],sort_results['sort_render_f3_lp_retest'],classhot_results['classhot_render_f3_srh_lp_first'],classhot_results['classhot_render_f3_srh_lp_retest'],classhot_results['classhot_render_f3_chk_lp_first'],classhot_results['classhot_render_f3_chk_lp_retest'],classcold_results['classcold_render_f3_chk_lp_first'],classcold_results['classcold_render_f3_chk_lp_retest']],
        [2400,'NA','NA',classhot_results['classhot_render_f4_srh_lp_first'],classhot_results['classhot_render_f4_srh_lp_retest'],'NA','NA','NA','NA'],
        [2900,sort_results['sort_render_f5_lp_first'],sort_results['sort_render_f5_lp_retest'],classhot_results['classhot_render_f5_srh_lp_first'],classhot_results['classhot_render_f5_srh_lp_retest'],classhot_results['classhot_render_f5_chk_lp_first'],classhot_results['classhot_render_f5_chk_lp_retest'],classcold_results['classcold_render_f5_chk_lp_first'],classcold_results['classcold_render_f5_chk_lp_retest']]
    )
    
    unit_df = pd.DataFrame(unit_table, columns=['SORT','CLASSHOT','CLASSCOLD'], index=['LOT','TP REV','BIN','SMP','BURN-IN'])
    de_df = pd.DataFrame(de_table, columns=['BEGIN','POST-HVQK','END_TRAD','END_CA','TPI_END','TPI_UPDT_END','BEGIN_TRAD','BEGIN_CA','TPI','TPI_UPT','BEGIN_TRAD','BEGIN_CA'], index=['DE_SKU @FIRST','DE_SKU @RETEST','','FEID @FIRST','FEID @RETEST'])
    gt_df = pd.DataFrame(gt_table, columns=['BEGIN','POST-HVQK','END_TRAD','END_CA','END_ISOL','TPI_END','TPI_UPDT_END','BEGIN_TRAD','BEGIN_CA','BEGIN_ISOL','TPI','TPI_UPDT','BEGIN_TRAD','BEGIN_CA','BEGIN_ISOL'], index=['GT_SKU @FIRST','GT_SKU @RETEST','','FRG @FIRST','FRG @RETEST','','FEID @FIRST','FEID @RETEST'])
    vmin_columns=['FREQ','FIRST','RETEST','FIRST','RETEST','FIRST','RETEST','FIRST','RETEST']
    media_df = pd.DataFrame(media_table, columns=vmin_columns, index=['F1','F2','F3','','F1','F2','F3'])
    render_df = pd.DataFrame(render_table, columns=vmin_columns, index=['F1','F2','F3','F4','F5','','F1','F2','F3','F4','F5'])
    #data = pd.concat([pd.Series(unit_results),pd.Series(sort_results),pd.Series(classhot_results), pd.Series(classcold_results)])
    #pd.set_option('display.max_rows',None)
    #print(data)
    
    return unit_df, de_df, gt_df, media_df, render_df

# Save filtered data to a new Excel file
def save_to_excel(df1, df2, df3, df4, df5, output_filepath, VID):
    
    try:
        wb = load_workbook(output_filepath)
        mode = 'a'
        sheet_exists = 'overlay'
    except:
        wb = None
        mode = 'w'
        sheet_exists = None
        
    with pd.ExcelWriter(output_filepath, engine='openpyxl', mode=mode, if_sheet_exists=sheet_exists) as writer:
        df1.to_excel(writer, sheet_name=VID, index=True, header=True, merge_cells=False, startrow=1, startcol=0)
        df2.to_excel(writer, sheet_name=VID, index=True, header=True, merge_cells=False, startrow=10, startcol=0)
        df3.to_excel(writer, sheet_name=VID, index=True, header=True, merge_cells=False, startrow=19, startcol=0)
        df4.to_excel(writer, sheet_name=VID, index=True, header=True, merge_cells=False, startrow=32, startcol=1)
        df5.to_excel(writer, sheet_name=VID, index=True, header=True, merge_cells=False, startrow=44, startcol=1)

def adjust_cells(output_file_path, width=18):
    # Load the workbook
    wb = load_workbook(output_file_path)

    # Loop through all sheets
    border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Adjust the width of all columns
        for col in ws.iter_cols():
            col_letter = col[0].column_letter  # Get column letter
            ws.column_dimensions[col_letter].width = width

        # Align all values in each cell and highlight its border
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
                if cell.value is not None:
                    cell.border = border_style
        
        # Insert Labels
        label_ranges = {
            'TABLE 1: UNIT DATA':'B1:D1',
            'TABLE 2: DE DATA':'B9:M9',
            'SEPERATOR_DE':'A14:M14',
            'TABLE 3: GT DATA':'B18:P18',
            'SEPERATOR_GT':'A23:P23',
            'SEPERATOR_GT2':'A26:P26',
            'TABLE 4: MEDIA VMIN DATA':'D30:K30',
            'SEPERATOR_MEDIA':'A37:K37',
            'TABLE 5: RENDER VMIN DATA':'D42:K42',
            'SEPERATOR_RENDER':'A51:K51',
            'UNIT_SORT':'B2',
            'UNIT_CLASSHOT':'C2',
            'UNIT_CLASSCOLD':'D2',
            'DE_SORT':'B10:G10',
            'DE_CLASSHOT':'H10:K10',
            'DE_CLASSCOLD':'L10:M10',
            'GT_SORT':'B19:H19',
            'GT_CLASSHOT':'I19:M19',
            'GT_CLASSCOLD':'N19:P19',
            'MEDIA_SORT':'D31:E31',
            'MEDIA_CLASSHOT':'F31:I31',
            'MEDIA_CLASSCOLD':'J31:K31',
            'RENDER_SORT':'D43:E43',
            'RENDER_CLASSHOT':'F43:I43',
            'RENDER_CLASSCOLD':'J43:K43',
            'MEDIA_VMIN':'A34:A36',
            'MEDIA_LP':'A38:A40',
            'RENDER_VMIN':'A46:A50',
            'RENDER_LP':'A52:A56',
            'MEDIA_80FG_S':'D32:E32',
            'MEDIA_SRH_CH':'F32:G32',
            'MEDIA_CHK_CH':'H32:I32',
            'MEDIA_SRH_CC':'J32:K32',
            'RENDER_80FG_S':'D44:E44',
            'RENDER_SRH_CH':'F44:G44',
            'RENDER_CHK_CH':'H44:I44',
            'RENDER_SRH_CC':'J44:K44',
            }
        for label, label_range in label_ranges.items():
            ws.merge_cells(label_range)
            if 'TABLE' in label:
                ws[label_range.split(':')[0]].value = label
                ws[label_range.split(':')[0]].font = Font(bold=True, size=16, color="00FFFFFF")
                ws[label_range.split(':')[0]].fill = PatternFill(start_color="000000FF", fill_type="solid")
                ws[label_range.split(':')[0]].alignment =  Alignment(horizontal='left')
            elif 'SEPERATOR' in label:
                ws[label_range.split(':')[0]].fill = PatternFill(start_color="00003300", fill_type="solid")
            else:
                ws[label_range.split(':')[0]].value = label.split('_')[1]
                ws[label_range.split(':')[0]].font = Font(bold=True, size=14)
                ws[label_range.split(':')[0]].alignment =  Alignment(horizontal='center', vertical='center')
                ws[label_range.split(':')[0]].fill = PatternFill(start_color="00C0C0C0", fill_type="solid")
                ws[label_range.split(':')[0]].border = border_style
                if 'SORT' in label:
                    ws[label_range.split(':')[0]].fill = PatternFill(start_color="00FFCC00", fill_type="solid")
                elif 'CLASSHOT' in label:
                    ws[label_range.split(':')[0]].fill = PatternFill(start_color="00FF6600", fill_type="solid")
                elif 'CLASSCOLD' in label:
                    ws[label_range.split(':')[0]].fill = PatternFill(start_color="0033CCCC", fill_type="solid")  
        
        # Compare results and highlight for any mismatch    
        result_compare_list = {
            'de_sku':['B12','C12','D12','E12','F12','F12','G12','H12','I12','J12','K12','L12','M12','B13','C13','D13','E13','F13','F13','G13','H13','I13','J13','K13','L13','M13'],
            'de_feid':['B15','C15','D15','E15','F15','F15','G15','H15','I15','J15','K15','L15','M15','B16','C16','D16','E16','F16','F16','G16','H16','I16','J16','K16','L16','M16'],
            'gt_sku':['B21','C21','D21','E21','F21','F21','G21','H21','I21','J21','K21','L21','M21','N21','O21','P21','B22','C22','D22','E22','F22','F22','G22','H22','I22','J22','K22','L22','M22','N22','O22','P22'],
            'gt_frg':['B24','C24','D24','E24','F24','F24','G24','H24','I24','J24','K24','L24','M24','N24','O24','P24','B25','C25','D25','E25','F25','F25','G25','H25','I25','J25','K25','L25','M25','N25','O25','P25'],
            'gt_feid':['B27','C27','D27','E27','F27','F27','G27','H27','I27','J27','K27','L27','M27','N27','O27','P27','B28','C28','D28','E28','F28','F28','G28','H28','I28','J28','K28','L28','M28','N28','O28','P28'],
        }
        for key, cell_list in result_compare_list.items():
            # Remove blanks
            cell_value_map = {cell:ws[cell].value for cell in cell_list if ws[cell].value != '-'}
            # Remain Unique Values
            duplicates = Counter(cell_value_map.values())
            if duplicates:
                max_occurs = max(duplicates.values())
                unique_cells = {cell:value for cell, value in cell_value_map.items() if duplicates[value] != max_occurs}
                for cell in unique_cells:
                    ws[cell].fill = PatternFill(start_color="00FFFF00", fill_type="solid")

    # Save the modified workbook
    wb.save(output_file_path)

def main(input_filepath, input_datatopull, output_filepath, vid_list, locn_list):
    # Load all Input Files
    rawdata_df = load_excel(input_filepath).fillna("")
    sort_df = pd.read_excel(input_datatopull, sheet_name="Sort")
    class_df = pd.read_excel(input_datatopull, sheet_name="Class")
    # Converting column values to list for later operation
    columns_to_list = ['Keywords', 'Exclude_Keywords']
    sort_df[columns_to_list] = sort_df[columns_to_list].map(lambda x: x.rsplit(",") if isinstance(x,str) else None)
    class_df[columns_to_list] = class_df[columns_to_list].map(lambda x: x.rsplit(",") if isinstance(x,str) else None)
    # Convert dataframe to dictionary
    sort_info = sort_df.set_index("Variable").to_dict(orient="index")
    class_info = class_df.set_index("Variable").to_dict(orient="index")
    # Loop through each unit for data pulling
    for unit in range(len(vid_list)):
        df1, df2, df3, df4, df5 = data_compile(rawdata_df, vid_list[unit], sort_info, class_info)
        save_to_excel(df1, df2, df3, df4, df5, output_filepath, vid_list[unit])
        print(f"Finished Extracting Data for VID [{vid_list[unit]}]")
    print("Data for All Units Extracted Successfully...\n")

    adjust_cells(output_filepath)
    print(f"Cells highlighted, fitted and saved to '{output_filepath}'\n")
    
if __name__ == "__main__":

    #'''
    parser = argparse.ArgumentParser(description='This script extracts specified data from input csv file for given VIDs')
    parser.add_argument("--inputfile", type=str, required=True, help="Input file name (include .csv)")
    parser.add_argument("--outputfile", type=str, required=True, help="Output file name (include .xlsx)")
    parser.add_argument("--vid", type=str, required=True, help='List of VIDs seperated by commas, ex. "U4F771Q300179,U4F771Q300180"')

    args = parser.parse_args()

    directory = str(Path(__file__).parent)
    input_filepath = Path('/'.join([directory, args.inputfile]))
    output_filepath = Path('/'.join([directory, args.outputfile]))
    vid_list = args.vid.split(",")
    
    print("="*100)
    print(f"INPUT FILE PATH: \t{input_filepath}")
    print(f"OUTPUT FILE PATH:\t{output_filepath}")
    print(f"VID List: \t\t{vid_list}")
    print("="*100)
    #'''
    '''
    # Input and Output file paths
    input_filepath = Path("C:\\PythonScripts\\TestData\\bin99.csv")  
    output_filepath = Path("C:\\PythonScripts\\TestData\\bin99_processed.xlsx")
    vid_list = ["U50MU73100314","U5ST331600407","U53Y6J8700683","U53Y6J8700669","U53Y6J8700222","U53Y6J8700276","U53Y6J8700227"]
    '''
    input_datatopull = Path("C:\\PythonScripts\\datatopull.xlsx") 
    locn_list = ["119325","6261","6212"]
    
    if output_filepath.is_file():
        output_filepath.unlink()
        print("Output file found in provided path, File Removed...\n")
    
    if not (input_filepath.exists() and input_datatopull.exists()) :
        print("Input file not found, aborting script...\n")
    else: 
        main(input_filepath, input_datatopull, output_filepath, vid_list, locn_list)
